from openpyxl import Workbook
import datetime
import time
from openpyxl.styles import Alignment

wb = Workbook() # 创建文件对象


ws = wb.active  # 获取第一个sheet

ws.cell(1,1, 'Test').alignment = Alignment(horizontal='center', vertical='center')


# ws['A2'] = datetime.datetime.now() # 写入一个当前时间
#
# # 写入一个自定义的时间格式
# # locale.setlocale(locale.LC_CTYPE, 'chinese')
# ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())
# ws['A4'] = 'Test'

# Save the file
wb.save("/home/banruo/chatPy.xlsx")