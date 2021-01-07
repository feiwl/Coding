from openpyxl.styles import Color, Font, Alignment
import openpyxl
import paramiko
import time
import re

# SSH connection object
class _DConnection:
    def __init__(self, host, port, username, password):
        self.ssh = paramiko.SSHClient()
        self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.ssh.connect(hostname=host, port=port, username=username, password=password)

    def execute(self, command):
        stdin, stdout, stderr = self.ssh.exec_command(command)
        if stderr.read().decode():
            raise AttributeError(stderr)
        return stdout.read().decode()

    def close(self):
        if self.ssh.get_transport():
            self.ssh.close()

class Get_SystemState:
    def __init__(self):
        self._memory_state = list()
        self._cpu_state = list()
        self._dick_current_state = list()
        self._global_config = dict()

    def get_host_cpu_state(self, conn) -> str:
        command = 'cat /proc/stat | grep "cpu "'

        std_out = conn.execute(command)
        cpu_time_list = re.findall('\d+', std_out)
        cpu_idle1 = cpu_time_list[3]
        total_cpu_time1 = 0
        for t in cpu_time_list:
            total_cpu_time1 = total_cpu_time1 + int(t)

        time.sleep(2)

        std_out = conn.execute(command)
        cpu_time_list = re.findall('\d+', std_out)
        cpu_idle2 = cpu_time_list[3]
        total_cpu_time2 = 0
        for t in cpu_time_list:
            total_cpu_time2 = total_cpu_time2 + int(t)

        cpu_usage = str(round(1 - (float(cpu_idle2) - float(cpu_idle1)) / (total_cpu_time2 - total_cpu_time1), 2)) + '%'

        return cpu_usage

    def get_host_memory_state(self, conn) -> str:
        command = 'free'
        std_out = conn.execute(command)
        memory = re.search('Mem:.*?\n', std_out).group().split()
        total_mem = int(re.search('\d+', memory[1]).group())
        used_mem = int(re.search('\d+', memory[2]).group())
        mem_usage = str(round(used_mem / total_mem, 2)) + '%'
        return mem_usage

    def get_host_disk_current_date(self, conn) -> str:
        disk_current_line = None
        disk_current_state = conn.execute('df -hT')
        for line in disk_current_state.split('\n'):
            data = line.split().pop()
            if data == '/':
                disk_current_line = line
                break
        disk_current_state = re.search('(\d{2}|\d{3})%', disk_current_line).group()
        if disk_current_line:
            return disk_current_state


    def obtain_states(self, conn, date_frame:float=60):
        '''
        :param conn: _DConnection
        :param date_frame: 60 s
        :return:
        '''
        try:
            while True:
                self._memory_state.append(self.get_host_memory_state(conn))
                self._cpu_state.append(self.get_host_cpu_state(conn))
                self._dick_current_state.append(self.get_host_disk_current_date(conn))
                time.sleep(date_frame)
        finally:
            conn.close()

    @property
    def state(self):
        return {'memory':max(self._memory_state),
                'cpu': max(self._cpu_state),
                'disk': self._dick_current_state.pop()}

class InForMation_File_Handler:
    IP_FORMAT = '(?P<ip>([0-9]{1,3}\.){3}[0-9]{1,3})'
    PORT_FORMAT = '(?P<port>\d+)'
    USERNAME_FORMAT = '(?P<username>.*)'
    PASSWORD_FORMAT = '(?P<password>.*)'
    INFO_FORMAT = re.compile('{IP=' + IP_FORMAT + ', PORT=' + PORT_FORMAT + ', USERNAME=' + USERNAME_FORMAT + ', PASSWORD=' + PASSWORD_FORMAT + '}')

    def __init__(self):
        self._system_info = list()
        self._global_config = dict()

    def infor_match_proc(self, infomation_file):
        with open(infomation_file, 'r') as file:
            for line in file:
                info = self.INFO_FORMAT.match(line)
                if info:
                    self._system_info.append({'ip': info.group('ip'), 'port': info.group('port'),
                                        'username': info.group('username'), 'password': info.group('password')})

                watch_date = re.search('WATCH_DATE=: .*', line)
                if watch_date:
                    watch_date = watch_date.group().split(':')[1].strip()
                    self._global_config['watch_date'] = int(watch_date)

                date_frame = re.search('DATE_FRAME=: \d+', line)
                if date_frame:
                    date_frame = date_frame.group().split(':')[1].strip()
                    self._global_config['date_frame'] = float(date_frame)

        if self._global_config:
            if self._global_config['date_frame']: pass
            if self._global_config['watch_date']: pass
        else:
            raise ValueError('Infomation_file <WATCH_DATE> ERROR...')

        if self._system_info:
            pass
        else:
            raise ValueError('Infomation_file <IP, PORT, USERNAME, PASSWD> ERROR...')

    @property
    def computer_infors(self):
        '''
        :return: self._system_info is Computer list < [{ip,port,username,password}, {...}, {...}] >
        '''
        return self._system_info

    @property
    def global_config(self):
        '''
        :return: self._global_config is dict {'begin_date': value, 'end_date': value, 'date_frame': value}
        :obtion: cls.global_config['key']
        '''
        return self._global_config


class openpyxl_handler:
    def __init__(self, to_excel_path):
        # Excel 初始化
        self.to_excel_path = to_excel_path
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

        # IP
        self.ws['A1'] = 'IP地址'
        self.ws.column_dimensions['A'].width = 15.0
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # CPU
        self.ws['B1'] = 'CPU 使用率'
        self.ws.column_dimensions['B'].width = 15.0
        self.ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

        # Memory
        self.ws['C1'] = 'Memory 使用率'
        self.ws.column_dimensions['C'].width = 15.0
        self.ws['C1'].alignment = Alignment(horizontal='center', vertical='center')

        # Disk
        self.ws['D1'] = 'Disk 使用率'
        self.ws.column_dimensions['D'].width = 15.0
        self.ws['D1'].alignment = Alignment(horizontal='center', vertical='center')

    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    def insert_excel_file(self, infor):
        """
        :param infor: (ip, cpu, memory, disk)
        :return:
        """
        current_row = self.getRowsClosNum()[0]+1

        # ip address
        self.ws['A{}'.format(current_row)] = infor[0]
        self.ws['A{}'.format(current_row)].alignment = Alignment(horizontal='center', vertical='center')

        # cpu rate
        self.ws['B{}'.format(current_row)] = infor[1]
        self.ws['B{}'.format(current_row)].alignment = Alignment(horizontal='center', vertical='center')

        # memory rate
        self.ws['C{}'.format(current_row)] = infor[2]
        self.ws['C{}'.format(current_row)].alignment = Alignment(horizontal='center', vertical='center')

        # disk rate
        self.ws['D{}'.format(current_row)] = infor[3]
        self.ws['D{}'.format(current_row)].alignment = Alignment(horizontal='center', vertical='center')

    def save(self):
        self.wb.save(self.to_excel_path)
        self.wb.close()

def time_control(watch_date:int):
    """
    :param watch_date: %H float
    :return:
    """
    if watch_date != abs(watch_date):
        raise ValueError('Parameter error < watch_date={} >'.format(watch_date))

    watch_date = watch_date * 3600
    time.sleep(watch_date)

conn = _DConnection('192.168.10.68', 22, 'feiwl', 'feiwenlong')
system = Get_SystemState()
print(system.get_host_disk_current_date(conn))
